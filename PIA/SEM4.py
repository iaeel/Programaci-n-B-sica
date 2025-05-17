import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
import re
from datetime import datetime
import os
import numpy as np
import io
from PIL import Image

class AnimeExcelGenerator:
    def __init__(self, json_file):
        self.json_file = json_file
        self.data = self.load_and_validate_data()
        
    def load_and_validate_data(self):
        """Carga y valida los datos del JSON"""
        try:
            with open(self.json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
            if not isinstance(data, list):
                data = [data]
                
            validated_data = []
            for item in data:
                # Validación con expresiones regulares
                valid_item = {
                    'titulo': self.validate_field(item, 'titulo', r'^[\w\sáéíóúñÁÉÍÓÚÑ¡!¿?.,-]+$', 'Desconocido'),
                    'puntuacion': self.validate_score(item.get('puntuacion', 0)),
                    'generos': self.validate_field(item, 'generos', r'^[\w\s,áéíóúñÁÉÍÓÚÑ]+$', 'Desconocido'),
                    'fecha_inicio': self.validate_year(item.get('fecha_inicio', '2000'))
                }
                validated_data.append(valid_item)
                
            return validated_data
            
        except Exception as e:
            raise Exception(f"Error al cargar datos: {str(e)}")
    
    def validate_field(self, item, field, regex, default):
        """Valida un campo con expresión regular"""
        value = str(item.get(field, default))
        return value if re.match(regex, value) else default
    
    def validate_score(self, score):
        """Valida que la puntuación esté entre 0 y 10"""
        try:
            score = float(score)
            return max(0, min(10, score))
        except:
            return 5.0  # Valor por defecto
    
    def validate_year(self, year_str):
        """Valida y extrae el año de la fecha"""
        try:
            if re.match(r'^\d{4}$', str(year_str)):
                year = int(year_str)
            elif re.match(r'^\d{4}-\d{2}-\d{2}$', str(year_str)):
                year = int(year_str.split('-')[0])
            else:
                year = 2000
                
            current_year = datetime.now().year
            return max(1900, min(current_year + 1, year))
        except:
            return 2000
    
    def generate_excel(self):
        """Genera el archivo Excel con gráficas incorporadas"""
        try:
            wb = openpyxl.Workbook()
            
            # ========== HOJA DE DATOS ==========
            ws_data = wb.active
            ws_data.title = "Datos Anime"
            headers = ["Título", "Puntuación", "Géneros", "Año", "Calificación"]
            ws_data.append(headers)
            
            # Formato de encabezados
            header_fill = PatternFill("solid", fgColor="4472C4")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col in range(1, len(headers)+1):
                cell = ws_data.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Llenar datos con formato condicional
            green_fill = PatternFill("solid", fgColor="C6E0B4")
            yellow_fill = PatternFill("solid", fgColor="FFE699")
            red_fill = PatternFill("solid", fgColor="F8CBAD")
            
            for item in self.data:
                score = item['puntuacion']
                if score >= 8:
                    rating = "Excelente"
                    fill = green_fill
                elif score >= 6:
                    rating = "Buena"
                    fill = yellow_fill
                else:
                    rating = "Regular"
                    fill = red_fill
                
                row = [
                    item['titulo'],
                    score,
                    item['generos'],
                    item['fecha_inicio'],
                    rating
                ]
                ws_data.append(row)
                
                for col in range(1, len(row)+1):
                    ws_data.cell(row=ws_data.max_row, column=col).fill = fill
            
            # ========== HOJA DE GRÁFICAS ==========
            ws_charts = wb.create_sheet("Gráficas")
            
            # Generar y guardar cada gráfica
            self._generate_line_chart()
            img_line = ExcelImage(io.BytesIO(self._save_plot_to_buffer()))
            ws_charts.add_image(img_line, 'B2')
            
            self._generate_bar_chart()
            img_bar = ExcelImage(io.BytesIO(self._save_plot_to_buffer()))
            ws_charts.add_image(img_bar, 'B25')
            
            self._generate_scatter_chart()
            img_scatter = ExcelImage(io.BytesIO(self._save_plot_to_buffer()))
            ws_charts.add_image(img_scatter, 'L2')
            
            self._generate_pie_chart()
            img_pie = ExcelImage(io.BytesIO(self._save_plot_to_buffer()))
            ws_charts.add_image(img_pie, 'L25')
            
            # Ajustar anchos de columna
            for sheet in wb:
                for column in sheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Guardar archivo
            output_file = 'anime_analysis_with_charts.xlsx'
            wb.save(output_file)
            print(f"\n✅ Excel con gráficas generado: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"\n❌ Error al generar Excel: {str(e)}")
            return None
    
    def _save_plot_to_buffer(self):
        """Guarda el gráfico actual en un buffer de memoria"""
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        plt.close()
        buf.seek(0)
        return buf.read()
    
    def _generate_line_chart(self):
        """Gráfico de líneas: Evolución por año"""
        year_avg = {}
        for item in self.data:
            year = item['fecha_inicio']
            year_avg[year] = year_avg.get(year, []) + [item['puntuacion']]
        
        sorted_years = sorted(year_avg.keys())
        avg_scores = [np.mean(year_avg[y]) for y in sorted_years]
        
        plt.figure(figsize=(10, 5))
        plt.plot(sorted_years, avg_scores, 'o-', color='#1f77b4', linewidth=2)
        plt.title('Evolución de Puntuación Promedio', fontsize=12)
        plt.xlabel('Año')
        plt.ylabel('Puntuación Promedio (0-10)')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.xticks(rotation=45)
    
    def _generate_bar_chart(self):
        """Gráfico de barras: Top géneros"""
        genres = [g.strip() for item in self.data for g in item['generos'].split(',') if g.strip()]
        genre_counts = {}
        for genre in genres:
            genre_counts[genre] = genre_counts.get(genre, 0) + 1
        
        top_genres = sorted(genre_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        
        plt.figure(figsize=(10, 5))
        bars = plt.bar(
            [g[0] for g in top_genres],
            [g[1] for g in top_genres],
            color=['#2ca02c', '#9467bd', '#8c564b', '#e377c2', '#7f7f7f']
        )
        
        plt.title('Top 5 Géneros Más Comunes', fontsize=12)
        plt.xlabel('Género')
        plt.ylabel('Cantidad de Animes')
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{int(height)}',
                    ha='center', va='bottom')
    
    def _generate_scatter_chart(self):
        """Diagrama de dispersión: Año vs Puntuación"""
        years = [item['fecha_inicio'] for item in self.data]
        scores = [item['puntuacion'] for item in self.data]
        
        plt.figure(figsize=(10, 5))
        scatter = plt.scatter(
            years, scores,
            c=scores, cmap='RdYlGn',
            alpha=0.6, edgecolors='w', linewidth=0.5,
            s=[s*20 for s in scores]
        )
        
        plt.title('Relación Año vs Puntuación', fontsize=12)
        plt.xlabel('Año de Lanzamiento')
        plt.ylabel('Puntuación Individual')
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.colorbar(scatter, label='Puntuación')
        
        z = np.polyfit(years, scores, 1)
        p = np.poly1d(z)
        plt.plot(years, p(years), "r--", label='Tendencia')
        plt.legend()
    
    def _generate_pie_chart(self):
        """Gráfico de pastel: Distribución de puntuaciones"""
        scores = [item['puntuacion'] for item in self.data]
        score_dist = {
            'Excelente (9-10)': sum(1 for s in scores if s >= 9),
            'Bueno (7-8.9)': sum(1 for s in scores if 7 <= s < 9),
            'Regular (5-6.9)': sum(1 for s in scores if 5 <= s < 7),
            'Bajo (0-4.9)': sum(1 for s in scores if s < 5)
        }
        
        plt.figure(figsize=(10, 5))
        wedges, texts, autotexts = plt.pie(
            score_dist.values(),
            labels=score_dist.keys(),
            autopct='%1.1f%%',
            startangle=90,
            colors=['#2ca02c', '#98df8a', '#ffbb78', '#ff7f0e'],
            explode=(0.1, 0, 0, 0),
            shadow=True,
            textprops={'fontsize': 10}
        )
        
        plt.title('Distribución de Puntuaciones', fontsize=12)
        plt.legend(
            wedges, score_dist.keys(),
            title="Categorías",
            loc="center left",
            bbox_to_anchor=(1, 0.5)
        )

def main():
    print("\n" + "="*60)
    print("📊 GENERADOR DE EXCEL CON GRÁFICAS INCORPORADAS".center(60))
    print("="*60)
    
    # Buscar archivos JSON automáticamente
    json_files = [f for f in os.listdir() if f.endswith('.json')]
    
    if not json_files:
        print("\n⚠️ No se encontraron archivos JSON en el directorio actual")
        return
    
    print("\nArchivos JSON disponibles:")
    for i, f in enumerate(json_files, 1):
        print(f"{i}. {f}")
    
    try:
        selection = int(input("\nSeleccione el número del archivo a analizar: ")) - 1
        json_file = json_files[selection]
        
        print(f"\n🔍 Analizando archivo: {json_file}")
        analyzer = AnimeExcelGenerator(json_file)
        
        print("\n📊 Generando Excel con gráficas incorporadas...")
        excel_file = analyzer.generate_excel()
        
        if excel_file:
            print("\n" + "="*60)
            print(f"✅ ANÁLISIS COMPLETADO: {excel_file}".center(60))
            print("="*60)
            
            # Intentar abrir el archivo automáticamente
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(excel_file)
                elif os.name == 'posix':  # macOS/Linux
                    os.system(f'open "{excel_file}"' if sys.platform == 'darwin' else f'xdg-open "{excel_file}"')
            except:
                print("\nℹ️ El archivo se generó pero no pudo abrirse automáticamente")
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")

if __name__ == "__main__":
    import sys
    main()